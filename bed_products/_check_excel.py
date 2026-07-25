import openpyxl
import os

# Check multiple bed Excel files for "床架尺寸" field
bed_dir = r'G:\Trae工作文件\顾家产品库\bed_products'
for folder in sorted(os.listdir(bed_dir)):
    fp = os.path.join(bed_dir, folder)
    if not os.path.isdir(fp) or folder.startswith('~$'):
        continue
    ppt_dir = os.path.join(fp, 'PPT')
    if not os.path.isdir(ppt_dir):
        continue
    for f in os.listdir(ppt_dir):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            path = os.path.join(ppt_dir, f)
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                found = False
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        v = str(ws.cell(r, c).value or '')
                        if '床架尺寸' in v or '床架' in v:
                            found = True
                            break
                    if found:
                        break
                if found:
                    print(f"\n=== {folder} ===")
                    for r in range(1, ws.max_row + 1):
                        vals = []
                        for c in range(1, min(ws.max_column + 1, 8)):
                            v = ws.cell(r, c).value
                            if v is not None:
                                vals.append(f"C{c}: {str(v)[:150]}")
                        if vals:
                            print(f"  Row {r}: | ".join(vals))
                wb.close()
            except Exception as e:
                print(f"  {folder}: ERROR {e}")
            break
