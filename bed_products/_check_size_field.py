import openpyxl
import os
import re

bed_dir = r'G:\Trae工作文件\顾家产品库\bed_products'
print(f"{'产品':<20} {'床架尺寸数据':<60}")
print("-"*80)
for folder in sorted(os.listdir(bed_dir)):
    fp = os.path.join(bed_dir, folder)
    if not os.path.isdir(fp) or folder.startswith('~$'):
        continue
    ppt_dir = os.path.join(fp, 'PPT')
    if not os.path.isdir(ppt_dir):
        print(f"{folder:<20} {'无PPT文件夹':<60}")
        continue
    for f in os.listdir(ppt_dir):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            path = os.path.join(ppt_dir, f)
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                found_size = False
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        v = str(ws.cell(r, c).value or '')
                        if '床架尺寸' in v:
                            # Found header, get value from next column
                            val = str(ws.cell(r, c+1).value or '').strip()
                            if val and val != 'None':
                                # Parse 3-part dimensions
                                parts = re.split(r'[\s]+', val)
                                parsed = []
                                for p in parts:
                                    if p and re.match(r'\d+[*×xX]\d+[*×xX]\d+', p):
                                        parsed.append(p)
                                if parsed:
                                    print(f"{folder:<20} {' '.join(parsed):<60}")
                                else:
                                    print(f"{folder:<20} {val[:60]:<60}")
                                found_size = True
                            break
                    if found_size:
                        break
                if not found_size:
                    print(f"{folder:<20} {'无床架尺寸字段':<60}")
                wb.close()
            except Exception as e:
                print(f"{folder:<20} {'ERROR: '+str(e)[:40]:<60}")
            break
