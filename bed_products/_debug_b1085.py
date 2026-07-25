import openpyxl
import re

path = r'G:\Trae工作文件\顾家产品库\bed_products\JD.B1085PQ3\PPT\B1085PQ3产品一页纸话术V1.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

for r in range(1, ws.max_row + 1):
    vals = []
    for c in range(1, min(8, ws.max_column + 1)):
        v = ws.cell(r, c).value
        if v is not None:
            s = str(v)[:200].replace('\n', '\\n')
            vals.append(f"C{c}: {s}")
    if vals:
        print(f"Row {r}: {' | '.join(vals)}")

wb.close()
