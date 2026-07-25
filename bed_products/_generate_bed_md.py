"""床架品类专用 - 生成 bed_products 的 MD 文档（含价格数据）"""
import os, re, json
import openpyxl

BASE_DIR = r'G:\Trae工作文件\顾家产品库'
OUTPUT_DIR = os.path.join(BASE_DIR, 'markdown_db')
BED_DIR = os.path.join(BASE_DIR, 'bed_products')
JSON_PATH = os.path.join(OUTPUT_DIR, '_price_data.json')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 加载价格数据
price_data = {'bed_frames': []}
if os.path.exists(JSON_PATH):
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        price_data = json.load(f)

frame_by_code = {}
for item in price_data['bed_frames']:
    frame_by_code.setdefault(item['货号'], []).append(item)

def normalize(s): return re.sub(r'[\s.\n]', '', s.upper())

def match_code(folder, code_dict):
    fn = normalize(folder)
    for code, items in code_dict.items():
        if normalize(code) == fn: return items
    fs = re.sub(r'^(JD|HS|BY)', '', fn)
    for code, items in code_dict.items():
        cs = re.sub(r'^(JD|B|90|HS|BY)', '', normalize(code))
        if fs and fs == cs: return items
        for n in re.findall(r'\d+[A-Z0-9]*', fs):
            if len(n) >= 4 and n in cs: return items
    return []

def safe_name(f): return f.replace(' ', '_').replace('/', '_')
def write_md(p, l):
    with open(p, 'w', encoding='utf-8') as f: f.write('\n'.join(l))
def price_str(v):
    if v == 0: return ''
    return f"¥{v:,}" if isinstance(v, int) else f"¥{v:,.0f}"

def list_images(d):
    if not os.path.isdir(d): return []
    exts = ('.jpg','.jpeg','.png','.gif','.webp','.bmp')
    return sorted([f for f in os.listdir(d) if f.lower().endswith(exts)])

def get_images(fp):
    r = {}
    for s in ['场景图','浏览图','入户实景图','白底图']:
        imgs = list_images(os.path.join(fp, s))
        if imgs: r[s] = imgs
    return r

def fmt_imgs(d, fp):
    lines = []
    for s, imgs in sorted(d.items()):
        lines.append(f"\n### {s} ({len(imgs)}张)")
        for img in imgs[:10]:
            rel = os.path.relpath(os.path.join(fp, s, img), BASE_DIR).replace('\\','/')
            lines.append(f"![]({rel})")
        if len(imgs) > 10: lines.append(f"... 共{len(imgs)}张")
    return lines

def read_xlsx(fp):
    ppt_dir = os.path.join(fp, 'PPT')
    if not os.path.isdir(ppt_dir): return None
    fs = os.listdir(ppt_dir)
    xlsx = [f for f in fs if f.endswith('.xlsx') and not f.startswith('~$')]
    if not xlsx: return None
    path = os.path.join(ppt_dir, xlsx[0])
    
    info, dims = {}, {'床架尺寸':'', '在售规格':''}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for r in range(1, ws.max_row+1):
            c1 = str(ws.cell(r,1).value or '').replace('\n','').strip()
            c2 = str(ws.cell(r,2).value or '').strip()
            c3 = str(ws.cell(r,3).value or '').strip()
            c4 = str(ws.cell(r,4).value or '').strip()
            c5 = str(ws.cell(r,5).value or '').strip()
            
            if '设计风格' in c2: info['设计风格'] = c3[:200]
            elif '产品配色' in c2 or '配色' in c2: info['产品配色'] = c3[:200]
            elif ('材质' in c2 or '面料' in c2 or '靠包' in c2) and '新材质' not in c1:
                if c3: info['材质'] = c3[:300].split('\n')[0].strip()
            elif ('一句话' in c2 or '卖点' in c1 or '核心' in c1) and '新材质' not in c1:
                if c3 and len(c3) > 5: info['卖点'] = c3[:500]
            
            # 床架尺寸在 D 列(Excel列4)，值在 E 列(Excel列5)
            if '床架尺寸' in c4:
                dims['床架尺寸'] = (c5 or '').replace('\\','\n').strip()
            if '在售规格' in c2:
                dims['在售规格'] = (c3 or '').replace('\\','\n').strip()
        wb.close()
    except Exception as e:
        print(f'  [WARN] xlsx: {e}')
    return {'file': xlsx[0], 'info': info, 'dims': dims}

# ==================== 主流程：只处理 bed_products ====================
print(">>> 床架品类 MD 生成")
count = 0
for folder in sorted(os.listdir(BED_DIR)):
    fp = os.path.join(BED_DIR, folder)
    if not os.path.isdir(fp) or folder.startswith(('~$', '__')): continue
    
    doc = read_xlsx(fp)
    images = get_images(fp)
    frames = match_code(folder, frame_by_code)
    
    lines = [f"# {folder}\n"]
    
    # 产品信息 + 床架尺寸
    # 解析床架尺寸(3值: 长度*宽度*床头高度)与在售规格的对应关系
    outer_sizes = []  # [(length, width, head_height), ...]
    sale_specs = []   # ["181*201", ...]
    
    if doc:
        info, dims = doc['info'], doc['dims']
        
        # 解析床架尺寸
        raw_sizes = dims.get('床架尺寸', '')
        if raw_sizes:
            for s in re.split(r'[\s\n]+', raw_sizes.strip()):
                s = s.strip()
                if not s: continue
                # 处理带|后缀的 (如 "216*185*115|气动杆")
                clean_s = s.split('|')[0].strip()
                # 处理 "2350/2260*163*116" 格式: 长度有/分隔的替代值,取第一个
                # 正则: 第一个数字, 可选/数字, 然后 *宽度 *高度
                m = re.match(r'(\d+)(?:/\d+)?[*×xX](\d+)[*×xX](\d+)', clean_s)
                if m:
                    try:
                        vals = [int(m.group(1)), int(m.group(2)), int(m.group(3))]
                        # 如果数值大于500, 说明是mm单位, 转成cm
                        vals = [v // 10 if v > 500 else v for v in vals]
                        outer_sizes.append(tuple(vals))
                    except ValueError:
                        pass
        
        # 解析在售规格 (床宽*床长 格式, 可能带"矮脚"等后缀)
        raw_specs = dims.get('在售规格', '')
        if raw_specs:
            raw_specs = raw_specs.replace('\\', '\n')
            for s in re.split(r'[\s\n]+', raw_specs.strip()):
                s = s.strip()
                if not s: continue
                # 保留完整spec(含后缀如"矮脚"),只验证包含数字x数字的核心格式
                if re.search(r'\d+[*×xX]\d+', s):
                    sale_specs.append(s)
        
        # 构建尺寸映射表: 在售规格(宽度×长度) → 床架外尺寸(长度×宽度×高度)
        # 后缀(如"矮脚")保留在key中,避免标准款与矮脚款互相覆盖
        size_map = {}  # { "201x181": "223×192×117", "201x181矮脚": "223×192×108", ... }
        for i, spec in enumerate(sale_specs):
            if i < len(outer_sizes):
                outer = outer_sizes[i]
                outer_str = f"{outer[0]}×{outer[1]}×{outer[2]}"
                # 提取数字部分和后缀
                nums = re.findall(r'\d+', spec)
                suffix = re.sub(r'[\d*×xX\s]', '', spec)
                if len(nums) >= 2:
                    # 原序: 宽度x长度+后缀 (在售规格格式)
                    key_orig = f"{nums[0]}x{nums[1]}{suffix}"
                    size_map[key_orig] = outer_str
                    # 反序: 长度x宽度+后缀 (价格表格式)
                    key_rev = f"{nums[1]}x{nums[0]}{suffix}"
                    size_map[key_rev] = outer_str
        
        # 输出产品信息
        lines.append("## 产品信息\n")
        for k in ['设计风格','产品配色','材质']:
            if info.get(k): lines.append(f"- **{k}**: {info[k]}")
        if info.get('卖点'): lines.append(f"\n### 核心卖点\n{info['卖点']}")
        if dims['在售规格']:
            lines.append(f"\n**在售规格**: {dims['在售规格']}")
        lines.append("")
    
    # 价格数据
    if frames:
        lines.append("## 床架\n")
        lines.append("> 内径规格格式: **宽度×长度** (床宽 × 床头到床尾距离)。如需短一点的床(床头到床尾短)，请选择长度数值较小的规格。\n")
        lines.append("> 床架外尺寸 = 总长度(床头到床尾) × 总宽度 × 床头高度，单位CM\n")
        
        by_color = {}
        for item in frames:
            by_color.setdefault(item.get('配色型号',''), []).append(item)
        for color, items in by_color.items():
            if color: lines.append(f"### 配色型号: {color}\n")
            # 解析价格表中规范的格式 (如 "201x181" 长度x宽度)
            spec_display = []
            for item in items:
                spec = item['规格']
                # 提取数字和后缀,构造匹配key
                nums = re.findall(r'\d+', spec)
                suffix = re.sub(r'[\d*xX*×\s]', '', spec)
                key = ''
                if len(nums) >= 2:
                    key = f"{nums[0]}x{nums[1]}{suffix}"
                outer = size_map.get(key, '')
                # 尝试非严格匹配(去后缀再试)
                if not outer and len(nums) >= 2:
                    key2 = f"{nums[0]}x{nums[1]}"
                    outer = size_map.get(key2, '')
                if outer:
                    spec_display.append(f"| {spec} | {outer} | {price_str(item['实际成交价'])} | {item.get('可搭配床垫厚度','')} |")
                else:
                    spec_display.append(f"| {spec} | — | {price_str(item['实际成交价'])} | {item.get('可搭配床垫厚度','')} |")
            
            if spec_display:
                lines.append("| 规格(内径) | 床架外尺寸(总长×总宽×头高) | 实际成交价 | 可搭配床垫厚度 |")
                lines.append("|-----------|---------------------------|-----------|--------------|")
                lines.extend(spec_display)
            lines.append("")
    
    # 图片
    if images:
        lines.append("## 图片素材\n")
        lines.extend(fmt_imgs(images, fp))
        lines.append("")
    
    write_md(os.path.join(OUTPUT_DIR, f"{safe_name(folder)}.md"), lines)
    ic = sum(len(v) for v in images.values())
    count += 1
    print(f"  {folder}: {len(frames)}床架 {ic}图")

print(f"床架品类完成: {count} 个文件")
