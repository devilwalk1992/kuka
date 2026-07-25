"""床垫品类专用 - 生成 mattress_products 的 MD 文档（含价格数据）"""
import os, re, json
import openpyxl

BASE_DIR = r'G:\Trae工作文件\顾家产品库'
OUTPUT_DIR = os.path.join(BASE_DIR, 'markdown_db')
MATTRESS_DIR = os.path.join(BASE_DIR, 'mattress_products')
JSON_PATH = os.path.join(OUTPUT_DIR, '_price_data.json')
os.makedirs(OUTPUT_DIR, exist_ok=True)

price_data = {'mattresses': []}
if os.path.exists(JSON_PATH):
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        price_data = json.load(f)

mattress_by_code = {}
for item in price_data['mattresses']:
    mattress_by_code.setdefault(item['货号'], []).append(item)

def normalize(s): return re.sub(r'[\s.\n]', '', s.upper())

def match_code(folder, code_dict):
    fn = normalize(folder)
    for code, items in code_dict.items():
        if normalize(code) == fn: return items
    fs = re.sub(r'^(JD|HS|BY)', '', fn)
    for code, items in code_dict.items():
        cs = re.sub(r'^(JD|B|90|HS|BY)', '', normalize(code))
        if fs and fs == cs: return items
        for n in re.findall(r'\d+[A-Z0-9]*', fs):
            if len(n) >= 4 and n in cs: return items
    return []

def safe_name(f): return f.replace(' ', '_').replace('/', '_')
def write_md(p, l):
    with open(p, 'w', encoding='utf-8') as f: f.write('\n'.join(l))
def price_str(v):
    if v == 0: return ''
    return f"¥{v:,}" if isinstance(v, int) else f"¥{v:,.0f}"

def list_images(d):
    if not os.path.isdir(d): return []
    exts = ('.jpg','.jpeg','.png','.gif','.webp','.bmp')
    return sorted([f for f in os.listdir(d) if f.lower().endswith(exts)])

def get_images(fp):
    r = {}
    for s in ['场景图','浏览图','入户实景图','白底图']:
        imgs = list_images(os.path.join(fp, s))
        if imgs: r[s] = imgs
    return r

def fmt_imgs(d, fp):
    lines = []
    for s, imgs in sorted(d.items()):
        lines.append(f"\n### {s} ({len(imgs)}张)")
        for img in imgs[:10]:
            rel = os.path.relpath(os.path.join(fp, s, img), BASE_DIR).replace('\\','/')
            lines.append(f"![]({rel})")
        if len(imgs) > 10: lines.append(f"... 共{len(imgs)}张")
    return lines

def read_xlsx(fp):
    ppt_dir = os.path.join(fp, 'PPT')
    if not os.path.isdir(ppt_dir): return None
    fs = os.listdir(ppt_dir)
    xlsx = [f for f in fs if f.endswith('.xlsx') and not f.startswith('~$')]
    if not xlsx: return None
    path = os.path.join(ppt_dir, xlsx[0])
    
    info = {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for r in range(1, ws.max_row+1):
            c1 = str(ws.cell(r,1).value or '').replace('\n','').strip()
            c2 = str(ws.cell(r,2).value or '').strip()
            c3 = str(ws.cell(r,3).value or '').strip()
            if not c2 or c2.startswith('='): continue
            if ('材质' in c2 or '面料' in c2) and '新材质' not in c1:
                if c3: info['材质'] = c3[:300].split('\n')[0].strip()
            elif '规格' in c2 or '尺寸' in c2:
                info['规格'] = c3[:300]
            elif '产品配置' in c2:
                info['产品配置'] = c3[:500]
            elif '睡感' in c2 and '等级' not in c2:
                info['睡感'] = c3[:300]
            elif ('一句话' in c2 or '卖点' in c1 or '核心' in c1) and '新材质' not in c1:
                if c3 and len(c3) > 5: info['卖点'] = c3[:500]
        wb.close()
    except Exception as e:
        print(f'  [WARN] xlsx: {e}')
    return {'file': xlsx[0], 'info': info}

# ==================== 主流程：只处理 mattress_products ====================
print(">>> 床垫品类 MD 生成")
count = 0
for folder in sorted(os.listdir(MATTRESS_DIR)):
    fp = os.path.join(MATTRESS_DIR, folder)
    if not os.path.isdir(fp) or folder.startswith('~$'): continue
    
    doc = read_xlsx(fp)
    images = get_images(fp)
    matched = match_code(folder, mattress_by_code)
    
    lines = [f"# {folder}\n"]
    
    if doc:
        lines.append("## 产品信息\n")
        info = doc['info']
        for k in ['材质','规格','产品配置','睡感']:
            if info.get(k): lines.append(f"- **{k}**: {info[k]}")
        if info.get('卖点'): lines.append(f"\n### 核心卖点\n{info['卖点']}")
        lines.append("")
    
    if matched:
        lines.append("## 床垫\n")
        s = matched[0].get('产品系列','')
        if s: lines.append(f"**产品系列**: {s}\n")
        m = matched[0].get('材质','')
        if m: lines.append(f"**材质**: {m}\n")
        lines.append("| 货号 | 规格 | 实际成交价 |")
        lines.append("|------|------|-----------|")
        for item in matched:
            lines.append(f"| {item['货号']} | {item['规格']} | {price_str(item['实际成交价'])} |")
        lines.append("")
    
    if images:
        lines.append("## 图片素材\n")
        lines.extend(fmt_imgs(images, fp))
        lines.append("")
    
    write_md(os.path.join(OUTPUT_DIR, f"{safe_name(folder)}.md"), lines)
    ic = sum(len(v) for v in images.values())
    count += 1
    print(f"  {folder}: {len(matched)}床垫 {ic}图")

print(f"床垫品类完成: {count} 个文件")
